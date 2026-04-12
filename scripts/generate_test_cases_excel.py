# -*- coding: utf-8 -*-

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_FILE = ROOT / "docs" / "test-case-muc-4.xlsx"
FALLBACK_OUTPUT_FILE = ROOT / "docs" / "test-case-muc-4-fixed.xlsx"


THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="F3E8FF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header(ws, row_index=1):
    for cell in ws[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_body(ws, start_row=2):
    for row in ws.iter_rows(
        min_row=start_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = LEFT
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autofit(ws, widths=None):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            line_length = max((len(line) for line in value.splitlines()), default=0)
            max_len = max(max_len, line_length)

        base_width = min(max(max_len + 2, 12), 45)
        ws.column_dimensions[letter].width = widths.get(col, base_width)


def add_table(ws, headers, rows, widths=None):
    ws.append(headers)
    style_header(ws, ws.max_row)

    for row in rows:
        ws.append(row)

    style_body(ws, 2)
    autofit(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook():
    wb = Workbook()

    # Sheet 1
    overview = wb.active
    overview.title = "Tong quan"
    overview["A1"] = "THÔNG TIN MỤC 4 - KIỂM THỬ THỦ CÔNG VÀ TỰ ĐỘNG"
    overview["A1"].font = Font(size=14, bold=True)
    overview["A1"].fill = SECTION_FILL
    overview["A1"].alignment = LEFT
    overview.merge_cells("A1:F1")

    overview.append(["Hạng mục", "Nội dung"])
    style_header(overview, 2)

    info_rows = [
        ("Tên hệ thống", "D&M Fashion & Apparel"),
        ("Link public", "https://dm-fashion-apparel.onrender.com"),
        ("Link local", "http://localhost:3000"),
        ("Tài khoản khách hàng", "customer@atelier.local / 123456"),
        ("Tài khoản quản trị", "admin@atelier.local / admin123"),
        ("Công cụ manual test", "Google Chrome"),
        ("Công cụ API test", "Postman"),
        ("Công cụ UI automation", "Selenium WebDriver"),
        (
            "Ghi chú",
            "Render bản miễn phí có thể sleep sau một thời gian không truy cập; "
            "dữ liệu runtime có thể bị reset sau khi redeploy.",
        ),
    ]

    for key, value in info_rows:
        overview.append([key, value])

    style_body(overview, 3)
    autofit(overview, {1: 24, 2: 90})
    overview.freeze_panes = "A3"

    # Sheet 2
    manual = wb.create_sheet("Kiem thu thu cong")
    manual_headers = [
        "STT",
        "Mã test case",
        "Nhóm chức năng",
        "Mức ưu tiên",
        "Điều kiện tiên quyết",
        "Bước thực hiện",
        "Dữ liệu kiểm thử",
        "Kết quả mong đợi",
        "Kết quả thực tế",
        "Trạng thái",
        "Ghi chú",
    ]
    manual_rows = [
        [1, "TC_UI_01", "Trang chủ", "Cao", "Website đang hoạt động", "1. Mở trang chủ\n2. Quan sát banner, menu, ô tìm kiếm, danh sách nổi bật", "URL trang chủ", "Trang chủ hiển thị đầy đủ, không lỗi bố cục, có menu và danh sách sản phẩm", "", "", ""],
        [2, "TC_UI_02", "Danh sách sản phẩm", "Cao", "Website đang hoạt động", "1. Mở trang /products\n2. Quan sát danh sách sản phẩm", "Không", "Danh sách sản phẩm hiển thị thành công, có ảnh, tên và giá", "", "", ""],
        [3, "TC_UI_03", "Tìm kiếm", "Cao", "Có dữ liệu sản phẩm", "1. Mở trang sản phẩm\n2. Nhập từ khóa vào ô tìm kiếm\n3. Gửi tìm kiếm", "Từ khóa: áo", "Hệ thống trả về các sản phẩm phù hợp với từ khóa", "", "", ""],
        [4, "TC_UI_04", "Tìm kiếm", "Trung bình", "Có dữ liệu sản phẩm", "1. Mở trang sản phẩm\n2. Nhập từ khóa không tồn tại\n3. Gửi tìm kiếm", "Từ khóa: xyz123", "Hệ thống không báo lỗi và hiển thị trạng thái không có kết quả phù hợp", "", "", ""],
        [5, "TC_UI_05", "Tìm kiếm", "Trung bình", "Có dữ liệu sản phẩm", "1. Mở trang sản phẩm\n2. Nhập từ khóa không dấu\n3. Gửi tìm kiếm", "Từ khóa: ao", "Hệ thống vẫn trả về các sản phẩm có tên chứa Áo", "", "", ""],
        [6, "TC_UI_06", "Lọc danh mục", "Trung bình", "Có dữ liệu sản phẩm", "1. Mở trang sản phẩm\n2. Chọn danh mục Nữ", "Danh mục: Nữ", "Chỉ các sản phẩm thuộc danh mục Nữ được hiển thị", "", "", ""],
        [7, "TC_UI_07", "Lọc danh mục", "Trung bình", "Có dữ liệu sản phẩm", "1. Mở trang sản phẩm\n2. Chọn danh mục Nam", "Danh mục: Nam", "Chỉ các sản phẩm thuộc danh mục Nam được hiển thị", "", "", ""],
        [8, "TC_UI_08", "Chi tiết sản phẩm", "Cao", "Có ít nhất 1 sản phẩm", "1. Mở danh sách sản phẩm\n2. Chọn một sản phẩm bất kỳ", "Sản phẩm hợp lệ", "Trang chi tiết hiển thị tên, giá, mô tả, ảnh và nút thêm vào giỏ hàng", "", "", ""],
        [9, "TC_UI_09", "Chi tiết sản phẩm", "Trung bình", "Sản phẩm có nhiều ảnh", "1. Mở chi tiết sản phẩm\n2. Quan sát ảnh chính và gallery", "Sản phẩm có gallery", "Ảnh sản phẩm hiển thị đúng, không vỡ và không lỗi đường dẫn", "", "", ""],
        [10, "TC_UI_10", "Giỏ hàng", "Cao", "Đang ở trang chi tiết sản phẩm", "1. Chọn số lượng hợp lệ\n2. Bấm Thêm vào giỏ hàng", "Số lượng: 1", "Sản phẩm được thêm vào giỏ hàng, số lượng giỏ hàng tăng", "", "", ""],
        [11, "TC_UI_11", "Giỏ hàng", "Cao", "Sản phẩm đã có trong giỏ hàng", "1. Thêm cùng một sản phẩm thêm một lần nữa", "Sản phẩm đã có sẵn trong giỏ", "Giỏ hàng gộp đúng sản phẩm và tăng số lượng", "", "", ""],
        [12, "TC_UI_12", "Giỏ hàng", "Cao", "Giỏ hàng đã có ít nhất 1 sản phẩm", "1. Mở giỏ hàng\n2. Thay đổi số lượng\n3. Bấm Cập nhật", "Số lượng mới: 2", "Giỏ hàng cập nhật đúng số lượng và thành tiền", "", "", ""],
        [13, "TC_UI_13", "Giỏ hàng", "Trung bình", "Giỏ hàng đã có ít nhất 1 sản phẩm", "1. Mở giỏ hàng\n2. Nhập số lượng bằng 0 hoặc âm\n3. Bấm Cập nhật", "Số lượng: 0", "Hệ thống xử lý phù hợp, sản phẩm bị xóa hoặc không chấp nhận giá trị không hợp lệ", "", "", ""],
        [14, "TC_UI_14", "Giỏ hàng", "Cao", "Giỏ hàng đã có ít nhất 1 sản phẩm", "1. Mở giỏ hàng\n2. Bấm Xóa sản phẩm", "Không", "Sản phẩm bị xóa khỏi giỏ hàng và tổng tiền được cập nhật", "", "", ""],
        [15, "TC_UI_15", "Giỏ hàng", "Trung bình", "Giỏ hàng trống", "1. Mở trang giỏ hàng", "Không", "Trang giỏ hàng hiển thị trạng thái trống rõ ràng, không phát sinh lỗi giao diện", "", "", ""],
        [16, "TC_UI_16", "Thanh toán", "Cao", "Giỏ hàng có sản phẩm", "1. Mở trang checkout\n2. Nhập đầy đủ thông tin giao hàng\n3. Gửi đơn", "Họ tên, email, số điện thoại, địa chỉ hợp lệ", "Đơn hàng được tạo thành công, hiển thị thông báo thành công", "", "", ""],
        [17, "TC_UI_17", "Thanh toán", "Cao", "Giỏ hàng có sản phẩm", "1. Mở trang checkout\n2. Bỏ trống một hoặc nhiều trường bắt buộc\n3. Gửi đơn", "Thiếu họ tên hoặc địa chỉ", "Hệ thống từ chối và thông báo cần điền đầy đủ thông tin", "", "", ""],
        [18, "TC_UI_18", "Thanh toán", "Trung bình", "Giỏ hàng trống", "1. Truy cập trực tiếp /checkout khi giỏ hàng chưa có sản phẩm", "Không", "Hệ thống chuyển hướng về trang sản phẩm hoặc cảnh báo giỏ hàng trống", "", "", ""],
        [19, "TC_UI_19", "Đăng nhập khách hàng", "Cao", "Website đang hoạt động", "1. Mở trang đăng nhập\n2. Nhập email và mật khẩu hợp lệ\n3. Bấm Đăng nhập", "customer@atelier.local / 123456", "Đăng nhập thành công và quay về trang phù hợp", "", "", ""],
        [20, "TC_UI_20", "Đăng nhập", "Cao", "Website đang hoạt động", "1. Mở trang đăng nhập\n2. Nhập sai mật khẩu\n3. Bấm Đăng nhập", "customer@atelier.local / sai-mat-khau", "Hệ thống từ chối đăng nhập và hiển thị thông báo lỗi", "", "", ""],
        [21, "TC_UI_21", "Đăng nhập", "Trung bình", "Website đang hoạt động", "1. Mở trang đăng nhập\n2. Bỏ trống email hoặc mật khẩu\n3. Gửi form", "Thiếu 1 trường bắt buộc", "Trình duyệt hoặc hệ thống yêu cầu nhập đủ trường bắt buộc", "", "", ""],
        [22, "TC_UI_22", "Đăng nhập quản trị", "Cao", "Website đang hoạt động", "1. Mở trang đăng nhập\n2. Nhập tài khoản admin\n3. Bấm Đăng nhập", "admin@atelier.local / admin123", "Đăng nhập thành công và chuyển đến trang quản trị", "", "", ""],
        [23, "TC_UI_23", "Đăng xuất", "Trung bình", "Đã đăng nhập", "1. Bấm Đăng xuất", "Không", "Phiên đăng nhập kết thúc và giao diện quay về trạng thái chưa đăng nhập", "", "", ""],
        [24, "TC_UI_24", "Phân quyền", "Trung bình", "Chưa đăng nhập admin", "1. Truy cập trực tiếp /admin", "Không", "Hệ thống không cho phép truy cập quản trị nếu không có quyền admin", "", "", ""],
        [25, "TC_UI_25", "Điều hướng", "Thấp", "Website đang hoạt động", "1. Bấm logo trên header từ một trang bất kỳ", "Không", "Người dùng được đưa về trang chủ", "", "", ""],
        [26, "TC_UI_26", "Trang lỗi", "Thấp", "Website đang hoạt động", "1. Truy cập một URL không tồn tại", "/abcxyz", "Hệ thống hiển thị trang 404 phù hợp, không lỗi server", "", "", ""],
    ]
    add_table(
        manual,
        manual_headers,
        manual_rows,
        {1: 8, 2: 16, 3: 18, 4: 12, 5: 28, 6: 40, 7: 28, 8: 42, 9: 28, 10: 12, 11: 18},
    )

    # Sheet 3
    api = wb.create_sheet("API Postman")
    api_headers = [
        "STT",
        "Mã test case",
        "Endpoint",
        "Phương thức",
        "Mục tiêu kiểm thử",
        "Điều kiện tiên quyết",
        "Headers",
        "Body / Tham số",
        "Kết quả mong đợi",
        "Kết quả thực tế",
        "Trạng thái",
        "Ghi chú",
    ]
    api_rows = [
        [1, "API_01", "/api/health", "GET", "Kiểm tra API health hoạt động", "Website đang hoạt động", "Không", "Không", "Trả về 200 và trạng thái ok", "", "", ""],
        [2, "API_02", "/api/products", "GET", "Lấy danh sách sản phẩm", "Website đang hoạt động", "Không", "Không", "Trả về 200, có trường total và danh sách products", "", "", ""],
        [3, "API_03", "/api/products?limit=3", "GET", "Kiểm tra giới hạn số lượng sản phẩm trả về", "Website đang hoạt động", "Không", "limit=3", "Trả về tối đa 3 sản phẩm", "", "", ""],
        [4, "API_04", "/api/products?q=ao", "GET", "Tìm kiếm sản phẩm không dấu", "Website đang hoạt động", "Không", "q=ao", "Trả về các sản phẩm liên quan đến áo", "", "", ""],
        [5, "API_05", "/api/products?category=Nữ", "GET", "Lọc sản phẩm theo danh mục", "Website đang hoạt động", "Không", "category=Nữ", "Chỉ trả về sản phẩm thuộc danh mục Nữ", "", "", ""],
        [6, "API_06", "/api/products/1", "GET", "Lấy chi tiết sản phẩm hợp lệ", "Sản phẩm ID 1 tồn tại", "Không", "Không", "Trả về 200 và thông tin sản phẩm ID 1", "", "", ""],
        [7, "API_07", "/api/products/9999", "GET", "Kiểm tra sản phẩm không tồn tại", "Website đang hoạt động", "Không", "Không", "Trả về 404 và message không tìm thấy sản phẩm", "", "", ""],
        [8, "API_08", "/api/auth/login", "POST", "Đăng nhập thành công với tài khoản khách hàng hợp lệ", "Website đang hoạt động", "Content-Type: application/json", '{"email":"customer@atelier.local","password":"123456"}', "Trả về 200, message thành công và object user", "", "", ""],
        [9, "API_09", "/api/auth/login", "POST", "Đăng nhập thành công với tài khoản admin hợp lệ", "Website đang hoạt động", "Content-Type: application/json", '{"email":"admin@atelier.local","password":"admin123"}', "Trả về 200, user có role admin", "", "", ""],
        [10, "API_10", "/api/auth/login", "POST", "Từ chối đăng nhập với mật khẩu sai", "Website đang hoạt động", "Content-Type: application/json", '{"email":"customer@atelier.local","password":"sai-mat-khau"}', "Trả về 401 và message thông tin đăng nhập không hợp lệ", "", "", ""],
        [11, "API_11", "/api/cart", "GET", "Kiểm tra giỏ hàng ban đầu", "Session mới, chưa thêm sản phẩm", "Không", "Không", "Trả về 200, itemCount bằng 0 hoặc danh sách items trống", "", "", ""],
        [12, "API_12", "/api/cart", "POST", "Thêm sản phẩm vào giỏ hàng qua API", "Sử dụng cùng session trong Postman", "Content-Type: application/json", '{"productId":1,"quantity":1}', "Trả về 201 và giỏ hàng có ít nhất 1 sản phẩm", "", "", ""],
        [13, "API_13", "/api/cart", "POST", "Thêm lại cùng sản phẩm vào giỏ hàng", "Session đã có sản phẩm ID 1 trong giỏ", "Content-Type: application/json", '{"productId":1,"quantity":1}', "Giỏ hàng tăng số lượng sản phẩm ID 1 thay vì tạo dòng mới sai lệch", "", "", ""],
        [14, "API_14", "/api/cart/1", "PATCH", "Cập nhật số lượng sản phẩm trong giỏ hàng", "Giỏ hàng đã có sản phẩm ID 1", "Content-Type: application/json", '{"quantity":2}', "Trả về 200 và số lượng sản phẩm được cập nhật thành 2", "", "", ""],
        [15, "API_15", "/api/cart/1", "PATCH", "Cập nhật số lượng về 0", "Giỏ hàng đã có sản phẩm ID 1", "Content-Type: application/json", '{"quantity":0}', "Sản phẩm bị xóa khỏi giỏ hàng hoặc giỏ hàng không còn sản phẩm ID 1", "", "", ""],
        [16, "API_16", "/api/cart/1", "DELETE", "Xóa sản phẩm khỏi giỏ hàng", "Giỏ hàng đã có sản phẩm ID 1", "Không", "Không", "Trả về 200 và giỏ hàng không còn sản phẩm ID 1", "", "", ""],
        [17, "API_17", "/api/orders", "POST", "Tạo đơn hàng hợp lệ", "Cần thêm ít nhất 1 sản phẩm vào giỏ hàng trước đó cùng session", "Content-Type: application/json", '{"customerName":"Nguyễn Văn A","customerEmail":"vana@example.com","customerPhone":"0901234567","shippingAddress":"123 Nguyễn Huệ, TP.HCM"}', "Trả về 201, có message tạo đơn hàng thành công và orderId", "", "", ""],
        [18, "API_18", "/api/orders", "POST", "Từ chối tạo đơn hàng khi giỏ hàng trống", "Session chưa có sản phẩm trong giỏ", "Content-Type: application/json", '{"customerName":"Nguyễn Văn A","customerEmail":"vana@example.com","customerPhone":"0901234567","shippingAddress":"123 Nguyễn Huệ, TP.HCM"}', "Trả về 400 và message giỏ hàng trống hoặc dữ liệu chưa hợp lệ", "", "", ""],
        [19, "API_19", "/api/orders", "POST", "Từ chối tạo đơn hàng khi thiếu dữ liệu", "Giỏ hàng có sản phẩm", "Content-Type: application/json", '{"customerName":"","customerEmail":"","customerPhone":"","shippingAddress":""}', "Trả về 400 và message dữ liệu thanh toán chưa hợp lệ", "", "", ""],
        [20, "API_20", "/api/admin/overview", "GET", "Từ chối truy cập admin khi chưa đăng nhập", "Chưa đăng nhập admin", "Không", "Không", "Trả về 403 và message cần quyền quản trị viên", "", "", ""],
        [21, "API_21", "/api/admin/overview", "GET", "Truy cập admin overview sau khi đăng nhập admin", "Đã gọi API login bằng tài khoản admin trong cùng session", "Không", "Không", "Trả về 200 và dữ liệu dashboard quản trị", "", "", ""],
    ]
    add_table(
        api,
        api_headers,
        api_rows,
        {1: 8, 2: 14, 3: 24, 4: 14, 5: 30, 6: 34, 7: 22, 8: 44, 9: 40, 10: 28, 11: 12, 12: 18},
    )

    # Sheet 4
    ui = wb.create_sheet("UI Selenium")
    ui_headers = [
        "STT",
        "Mã test case",
        "Luồng tự động hóa",
        "Mục tiêu",
        "Điều kiện tiên quyết",
        "Các bước tự động hóa",
        "Dữ liệu kiểm thử",
        "Kết quả mong đợi",
        "Công cụ / Ghi chú",
        "Trạng thái",
        "Ghi chú",
    ]
    ui_rows = [
        [1, "SEL_01", "Đăng nhập thành công", "Kiểm tra người dùng đăng nhập thành công trên giao diện", "Website hoạt động, trình duyệt sẵn sàng", "1. Mở trang /login\n2. Nhập email\n3. Nhập mật khẩu\n4. Bấm Đăng nhập\n5. Kiểm tra chuyển trang thành công", "customer@atelier.local / 123456", "Hệ thống đăng nhập thành công và hiển thị trạng thái phù hợp", "Selenium WebDriver + Chrome", "", ""],
        [2, "SEL_02", "Đăng nhập thất bại", "Kiểm tra thông báo lỗi khi nhập sai mật khẩu", "Website hoạt động, trình duyệt sẵn sàng", "1. Mở trang /login\n2. Nhập email hợp lệ\n3. Nhập mật khẩu sai\n4. Bấm Đăng nhập\n5. Kiểm tra thông báo lỗi", "customer@atelier.local / sai-mat-khau", "Hệ thống từ chối đăng nhập và hiển thị thông báo phù hợp", "Selenium WebDriver + Chrome", "", ""],
        [3, "SEL_03", "Tìm kiếm sản phẩm", "Kiểm tra ô tìm kiếm và kết quả hiển thị", "Website có dữ liệu sản phẩm", "1. Mở trang /products\n2. Nhập từ khóa áo\n3. Gửi tìm kiếm\n4. Kiểm tra danh sách kết quả", "Từ khóa: áo", "Danh sách trả về các sản phẩm phù hợp", "Selenium WebDriver + Chrome", "", ""],
        [4, "SEL_04", "Lọc danh mục", "Kiểm tra lọc sản phẩm theo danh mục", "Website có dữ liệu sản phẩm", "1. Mở trang /products\n2. Chọn danh mục Nữ\n3. Kiểm tra danh sách kết quả", "Danh mục: Nữ", "Kết quả hiển thị đúng theo danh mục đã chọn", "Selenium WebDriver + Chrome", "", ""],
        [5, "SEL_05", "Mở chi tiết sản phẩm", "Kiểm tra điều hướng từ danh sách sang chi tiết", "Có ít nhất 1 sản phẩm hiển thị", "1. Mở trang /products\n2. Chọn sản phẩm đầu tiên\n3. Kiểm tra trang chi tiết", "Sản phẩm hợp lệ", "Trang chi tiết hiển thị đúng tên sản phẩm và nút thêm vào giỏ hàng", "Selenium WebDriver + Chrome", "", ""],
        [6, "SEL_06", "Thêm vào giỏ hàng", "Kiểm tra luồng thêm sản phẩm vào giỏ hàng", "Có ít nhất 1 sản phẩm khả dụng", "1. Mở chi tiết sản phẩm\n2. Chọn số lượng\n3. Bấm Thêm vào giỏ hàng\n4. Mở giỏ hàng\n5. Kiểm tra sản phẩm đã có trong giỏ", "Sản phẩm hợp lệ", "Giỏ hàng hiển thị đúng sản phẩm vừa thêm", "Selenium WebDriver + Chrome", "", ""],
        [7, "SEL_07", "Cập nhật giỏ hàng", "Kiểm tra thay đổi số lượng trong giỏ hàng", "Giỏ hàng đã có ít nhất 1 sản phẩm", "1. Mở giỏ hàng\n2. Nhập số lượng mới\n3. Bấm Cập nhật\n4. Kiểm tra tổng tiền", "Số lượng mới: 2", "Số lượng và thành tiền được cập nhật đúng", "Selenium WebDriver + Chrome", "", ""],
        [8, "SEL_08", "Xóa sản phẩm khỏi giỏ hàng", "Kiểm tra nút xóa sản phẩm trong giỏ hàng", "Giỏ hàng đã có ít nhất 1 sản phẩm", "1. Mở giỏ hàng\n2. Bấm Xóa\n3. Kiểm tra giỏ hàng sau khi xóa", "Không", "Sản phẩm bị loại khỏi giỏ hàng thành công", "Selenium WebDriver + Chrome", "", ""],
        [9, "SEL_09", "Thanh toán thành công", "Kiểm tra luồng checkout cơ bản", "Giỏ hàng đã có sản phẩm", "1. Mở trang checkout\n2. Nhập thông tin giao hàng\n3. Gửi form\n4. Kiểm tra thông báo thành công", "Dữ liệu giao hàng hợp lệ", "Đơn hàng được tạo thành công", "Selenium WebDriver + Chrome", "", ""],
        [10, "SEL_10", "Phân quyền admin", "Kiểm tra người dùng chưa đăng nhập không vào được trang admin", "Chưa đăng nhập admin", "1. Truy cập trực tiếp /admin\n2. Kiểm tra nội dung trả về", "Không", "Trang quản trị không được truy cập nếu không có quyền admin", "Selenium WebDriver + Chrome", "", ""],
    ]
    add_table(
        ui,
        ui_headers,
        ui_rows,
        {1: 8, 2: 14, 3: 20, 4: 26, 5: 28, 6: 42, 7: 28, 8: 36, 9: 24, 10: 12, 11: 18},
    )

    # Sheet 5
    bug = wb.create_sheet("Bug report mau")
    bug_headers = [
        "STT",
        "Bug ID",
        "Mức độ",
        "Chức năng",
        "Tiêu đề lỗi",
        "Bước tái hiện",
        "Kết quả mong đợi",
        "Kết quả thực tế",
        "Trạng thái",
        "Người ghi nhận",
        "Ghi chú",
    ]
    bug_rows = [
        [
            1,
            "BUG_01",
            "Medium",
            "Tìm kiếm",
            "Ví dụ: Tìm kiếm không hiển thị kết quả đúng",
            "1. Mở trang sản phẩm\n2. Nhập từ khóa\n3. Gửi tìm kiếm",
            "Hiển thị kết quả phù hợp",
            "Kết quả trả về sai hoặc trống",
            "Open",
            "",
            "Điền khi phát sinh lỗi thật",
        ]
    ]
    add_table(
        bug,
        bug_headers,
        bug_rows,
        {1: 8, 2: 12, 3: 12, 4: 18, 5: 34, 6: 34, 7: 28, 8: 28, 9: 12, 10: 18, 11: 22},
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True
        for row_index in range(1, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 24
        ws.row_dimensions[1].height = 28

    return wb


if __name__ == "__main__":
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    try:
        workbook.save(OUTPUT_FILE)
        print(OUTPUT_FILE)
    except PermissionError:
        workbook.save(FALLBACK_OUTPUT_FILE)
        print(FALLBACK_OUTPUT_FILE)
